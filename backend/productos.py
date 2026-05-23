import openpyxl

def leer_productos(ruta):
    productos = []
    archivo = openpyxl.load_workbook(ruta)
    hoja = archivo.active

    for fila in hoja.iter_rows(min_row=2, values_only=True):
        productos.append({
            "id": fila[0],
            "nombre": fila[1],
            "precio": fila[2],
            "cantidad": fila[3]
        })

    return productos


def buscar_producto(productos, id_buscar):
    for p in productos:
        if p["id"] == id_buscar:
            return p
    return None
def crear_producto(ruta, id_producto, nombre, precio, cantidad):
    productos = leer_productos(ruta)

    for p in productos:
        if p["id"] == id_producto:
            return False

    archivo = openpyxl.load_workbook(ruta)
    hoja = archivo.active

    hoja.append([id_producto, nombre, precio, cantidad])

    archivo.save(ruta)

    return True
    def actualizar_producto(ruta, id_producto, nombre, precio, cantidad):
    archivo = openpyxl.load_workbook(ruta)
    hoja = archivo.active

    encontrado = False

    for fila in hoja.iter_rows(min_row=2):
        if fila[0].value == id_producto:
            fila[1].value = nombre
            fila[2].value = precio
            fila[3].value = cantidad
            encontrado = True
            break

    archivo.save(ruta)

    return encontrado
    def eliminar_producto(ruta, id_producto):
    archivo = openpyxl.load_workbook(ruta)
    hoja = archivo.active

    encontrado = False

    for i, fila in enumerate(hoja.iter_rows(min_row=2), start=2):
        if fila[0].value == id_producto:
            hoja.delete_rows(i)
            encontrado = True
            break

    archivo.save(ruta)

    return encontrado
